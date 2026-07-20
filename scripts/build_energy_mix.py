from __future__ import annotations

import json
import math
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SOURCE_FILE = ROOT / "各省装机与发电量.xlsx"
DATA_DIR = ROOT / "data"
OUTPUT_FILE = DATA_DIR / "energy-mix-data.json"
OUTPUT_JS_FILE = DATA_DIR / "energy-mix-data.js"

CAPACITY_FIELDS = [
    ("燃煤", "燃煤（万千瓦）"),
    ("水电", "水电(万千瓦)"),
    ("风电", "风电(万千瓦)"),
    ("光伏", "光伏(万千瓦)"),
    ("燃气", "燃气（万千瓦）"),
    ("核电", "核电(万千瓦)"),
    ("其他", "其他（生物质、储能）"),
]

GENERATION_FIELDS = [
    ("火电", "火电(亿千瓦时)"),
    ("水电", "水电(亿千瓦时)"),
    ("风电", "风电(亿千瓦时)"),
    ("光伏", "光伏(亿千瓦时)"),
    ("核电", "核电(亿千瓦时)"),
]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def number_value(value: Any) -> float:
    if value in (None, ""):
        return 0
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return 0
    if math.isnan(numeric) or math.isinf(numeric):
        return 0
    return round(numeric, 6)


def build_series(row: tuple[Any, ...], header_map: dict[str, int], fields: list[tuple[str, str]]) -> list[dict[str, Any]]:
    series = []
    for label, header in fields:
        index = header_map.get(header)
        value = number_value(row[index] if index is not None and index < len(row) else None)
        if value > 0:
            series.append({"name": label, "value": value})
    return series


def parse_year(sheet_name: str) -> str:
    digits = "".join(ch for ch in sheet_name if ch.isdigit())
    return digits[:4] if len(digits) >= 4 else sheet_name


def build_year_dataset(workbook: Any, sheet_name: str) -> dict[str, Any]:
    year = parse_year(sheet_name)
    ws = workbook[sheet_name]
    headers = [clean_text(value) for value in next(ws.iter_rows(min_row=2, max_row=2, values_only=True))]
    header_map = {header: index for index, header in enumerate(headers) if header}
    province_col = header_map["省份/电网区域"]

    provinces: dict[str, Any] = {}
    current_region = ""
    for row in ws.iter_rows(min_row=3, values_only=True):
        if clean_text(row[0]):
            current_region = clean_text(row[0])
        province = clean_text(row[province_col] if province_col < len(row) else None)
        if not province:
            continue
        capacity = build_series(row, header_map, CAPACITY_FIELDS)
        generation = build_series(row, header_map, GENERATION_FIELDS)
        provinces[province] = {
            "name": province,
            "region": current_region,
            "year": year,
            "capacityTotal": number_value(row[header_map["总装机容量(万千瓦)"]]),
            "generationTotal": number_value(row[header_map["总发电量(亿千瓦时)"]]),
            "capacity": capacity,
            "generation": generation,
        }
    return {"year": year, "sheet": sheet_name, "provinces": provinces}


def build_dataset() -> dict[str, Any]:
    workbook = openpyxl.load_workbook(SOURCE_FILE, read_only=True, data_only=True)
    year_items = [build_year_dataset(workbook, sheet_name) for sheet_name in workbook.sheetnames if "年" in sheet_name]
    years = {item["year"]: {"sheet": item["sheet"], "provinces": item["provinces"]} for item in year_items}
    year_order = sorted(years)

    return {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "source": {
            "file": SOURCE_FILE.name,
            "sheets": [item["sheet"] for item in year_items],
            "note": "由 scripts/build_energy_mix.py 根据各省装机与发电量.xlsx 生成。",
        },
        "units": {
            "capacity": "万千瓦",
            "generation": "亿千瓦时",
        },
        "defaultYear": year_order[-1] if year_order else "",
        "years": years,
    }


def main() -> None:
    DATA_DIR.mkdir(exist_ok=True)
    payload = build_dataset()
    json_text = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    OUTPUT_FILE.write_text(json_text, encoding="utf-8")
    OUTPUT_JS_FILE.write_text(f"window.ENERGY_MIX_DATA={json_text};\n", encoding="utf-8")
    count = sum(len(item["provinces"]) for item in payload["years"].values())
    print(f"wrote {OUTPUT_FILE.relative_to(ROOT)} and {OUTPUT_JS_FILE.relative_to(ROOT)} with {len(payload['years'])} years and {count} province-year records")


if __name__ == "__main__":
    main()
