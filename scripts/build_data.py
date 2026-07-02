from __future__ import annotations

import json
import math
import re
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
MAPPING_FILE = ROOT / "映射表字段.xlsx"
DATA_DIR = ROOT / "data"
OUTPUT_FILE = DATA_DIR / "dashboard-data.json"
OUTPUT_JS_FILE = DATA_DIR / "dashboard-data.js"

GROUP_LABELS = {
    "日前柱状图": "day_bar",
    "日前折线图": "day_line",
    "实时柱状图": "realtime_bar",
    "实时折线图": "realtime_line",
    "日前实时对比": "compare",
    "日前实时边界对比": "compare",
}

SECTION_HINTS = {
    "day": ("日前边界数据", "日前"),
    "realtime": ("实时边界数据", "实时"),
    "price": ("分区价格", "价格"),
    "day_price": ("日前", "分区价格", "价格"),
    "realtime_price": ("实时", "分区价格", "价格"),
}

FIELD_ALIASES = {
    "统一结算点电价-日前": ("统一结算点价格",),
    "日前平均出清价格": ("平均出清价格",),
    "实时平均出清价格": ("平均出清价格",),
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return round(value, 6)
    return value


def normalize_date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    text = clean_text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return text


def normalize_time(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, (int, float)):
        minutes = round(float(value) * 24 * 60)
        return f"{minutes // 60:02d}:{minutes % 60:02d}"
    text = clean_text(value)
    if not text:
        return None
    match = re.search(r"(\d{1,2}):(\d{2})", text)
    if match:
        return f"{int(match.group(1)):02d}:{match.group(2)}"
    return text


def read_mapping() -> dict[str, Any]:
    workbook = openpyxl.load_workbook(MAPPING_FILE, read_only=True, data_only=True)
    result: dict[str, Any] = {}
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        province = sheet_name.strip()
        config = {
            "day": {"barPrimary": None, "stackBars": [], "line": None},
            "realtime": {"barPrimary": None, "stackBars": [], "line": None},
            "compare": {
                "dayBarPrimary": None,
                "dayBarSecondary": None,
                "realtimeBarPrimary": None,
                "realtimeBarSecondary": None,
                "dayLine": None,
                "realtimeLine": None,
            },
        }
        current_group = ""
        for row in ws.iter_rows(values_only=True):
            first = clean_text(row[0] if row else "")
            if first in GROUP_LABELS:
                current_group = GROUP_LABELS[first]
                continue
            fields = [clean_text(value) for value in row[1:] if clean_text(value)]
            if not fields or not current_group:
                continue
            if current_group == "compare":
                if first == "日前柱子1":
                    config["compare"]["dayBarPrimary"] = fields[0]
                elif first == "日前柱子2":
                    config["compare"]["dayBarSecondary"] = fields[0]
                elif first == "实时柱子1":
                    config["compare"]["realtimeBarPrimary"] = fields[0]
                elif first == "实时柱子2":
                    config["compare"]["realtimeBarSecondary"] = fields[0]
                elif first in {"折线1", "日前折线1"}:
                    config["compare"]["dayLine"] = fields[0]
                elif first in {"折线2", "实时折线2"}:
                    config["compare"]["realtimeLine"] = fields[0]
                continue
            target = "day" if current_group.startswith("day") else "realtime"
            if first == "柱子1":
                config[target]["barPrimary"] = fields[0]
            elif first == "柱子2":
                config[target]["stackBars"] = fields
            elif first == "折线1":
                config[target]["line"] = fields[0]
        result[province] = config
    return result


def detect_header(ws: Any) -> tuple[int, list[str], list[str]]:
    best_row = 1
    best_score = -1
    best_values: list[str] = []
    section_labels = {"日前边界数据", "实时边界数据", "分区价格", "日前", "实时"}
    for row_index, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 8), values_only=True), start=1):
        values = [clean_text(value) for value in row]
        score = int("日期" in values) + int("时刻" in values) + sum(1 for value in values if value and value not in section_labels)
        if score > best_score:
            best_row = row_index
            best_score = score
            best_values = values

    section_row = [""] * len(best_values)
    if best_row > 1:
        raw_section = next(ws.iter_rows(min_row=best_row - 1, max_row=best_row - 1, values_only=True))
        section_row = [clean_text(value) for value in raw_section]
    return best_row, best_values, section_row


def choose_column(field: str, columns: list[dict[str, Any]], mode: str) -> int | None:
    candidates = (field, *FIELD_ALIASES.get(field, ()))
    exact = [column for column in columns if column["field"] in candidates]
    if not exact:
        return None
    hints = SECTION_HINTS.get(mode, ())
    hinted = [column for column in exact if any(hint in column["section"] for hint in hints)]
    if hinted:
        return hinted[0]["index"]
    if mode == "day":
        return exact[0]["index"]
    if mode == "realtime":
        return exact[-1]["index"]
    return exact[0]["index"]


def iter_workbook_records(path: Path, province_config: dict[str, Any]) -> tuple[list[dict[str, Any]], list[str]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = workbook[workbook.sheetnames[0]]
    header_row, headers, sections = detect_header(ws)
    columns = [
        {"index": index, "field": header, "section": sections[index] if index < len(sections) else ""}
        for index, header in enumerate(headers)
        if header
    ]
    date_col = choose_column("日期", columns, "day")
    time_col = choose_column("时刻", columns, "day")
    if date_col is None or time_col is None:
        return [], [f"{path.name}: 未找到日期或时刻列"]

    required_fields: dict[str, dict[str, str]] = {}
    for mode in ("day", "realtime"):
        cfg = province_config[mode]
        if cfg["barPrimary"]:
            required_fields[f"{mode}.barPrimary"] = {"field": cfg["barPrimary"], "type": mode}
        if cfg["line"]:
            required_fields[f"{mode}.line"] = {"field": cfg["line"], "type": f"{mode}_price"}
        for field in cfg["stackBars"]:
            required_fields[f"{mode}.stack.{field}"] = {"field": field, "type": mode}

    compare_cfg = province_config.get("compare", {})
    compare_requirements = {
        "compare.dayBarPrimary": ("dayBarPrimary", "day"),
        "compare.dayBarSecondary": ("dayBarSecondary", "day"),
        "compare.realtimeBarPrimary": ("realtimeBarPrimary", "realtime"),
        "compare.realtimeBarSecondary": ("realtimeBarSecondary", "realtime"),
        "compare.dayLine": ("dayLine", "day_price"),
        "compare.realtimeLine": ("realtimeLine", "realtime_price"),
    }
    for key, (config_key, value_type) in compare_requirements.items():
        field = compare_cfg.get(config_key)
        if field:
            required_fields[key] = {"field": field, "type": value_type}

    column_map: dict[str, int] = {}
    warnings: list[str] = []
    for key, meta in required_fields.items():
        field = meta["field"]
        value_type = meta["type"]
        index = choose_column(field, columns, value_type)
        if index is None:
            warnings.append(f"{path.name}: 字段缺失，已按 0 填充 - {field}")
        else:
            column_map[key] = index

    records: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        row_date = normalize_date(row[date_col] if date_col < len(row) else None)
        row_time = normalize_time(row[time_col] if time_col < len(row) else None)
        if not row_date or not row_time or row_date == "日期":
            continue
        values: dict[str, Any] = {}
        for key in required_fields:
            index = column_map.get(key)
            values[key] = 0 if index is None else json_value(row[index] if index < len(row) else None)
        records.append({"date": row_date, "time": row_time, "values": values})
    return records, warnings


def build_dataset() -> dict[str, Any]:
    mapping = read_mapping()
    provinces: dict[str, Any] = {}
    all_warnings: list[str] = []
    for province_dir in sorted(path for path in ROOT.iterdir() if path.is_dir() and not path.name.startswith(".")):
        province = province_dir.name
        if province not in mapping:
            continue
        records_by_time: dict[tuple[str, str], dict[str, Any]] = {}
        files = sorted(province_dir.glob("*.xlsx"))
        for workbook_path in files:
            workbook_records, warnings = iter_workbook_records(workbook_path, mapping[province])
            for record in workbook_records:
                key = (record["date"], record["time"])
                target = records_by_time.setdefault(key, {"date": record["date"], "time": record["time"], "values": {}})
                for value_key, value in record["values"].items():
                    current = target["values"].get(value_key)
                    if current in (None, 0) or value not in (None, 0):
                        target["values"][value_key] = value
            all_warnings.extend([f"{province}/{warning}" for warning in warnings])
        records = list(records_by_time.values())
        records.sort(key=lambda record: (record["date"], record["time"]))
        dates = sorted({record["date"] for record in records})
        provinces[province] = {
            "name": province,
            "files": [file.name for file in files],
            "dates": dates,
            "mapping": mapping[province],
            "records": records,
        }

    return {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "source": {
            "mappingFile": MAPPING_FILE.name,
            "note": "由 scripts/build_data.py 根据省份目录和映射表字段.xlsx 生成。",
        },
        "warnings": sorted(set(all_warnings)),
        "provinces": provinces,
    }


def main() -> None:
    DATA_DIR.mkdir(exist_ok=True)
    payload = build_dataset()
    json_text = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    OUTPUT_FILE.write_text(json_text, encoding="utf-8")
    OUTPUT_JS_FILE.write_text(f"window.DASHBOARD_DATA={json_text};\n", encoding="utf-8")
    count = sum(len(item["records"]) for item in payload["provinces"].values())
    print(f"wrote {OUTPUT_FILE.relative_to(ROOT)} and {OUTPUT_JS_FILE.relative_to(ROOT)} with {len(payload['provinces'])} provinces and {count} records")
    if payload["warnings"]:
        print("warnings:")
        for warning in payload["warnings"]:
            print(f"- {warning}")


if __name__ == "__main__":
    main()
